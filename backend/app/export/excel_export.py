import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
def invoices_to_xlsx_stream(invoices):
    wb = Workbook()
    ws = wb.active
    ws.append(["id","invoice_number","date","type","customer_name","currency","taxable_value","tax_amount","total","items"])
    for inv in invoices:
        ws.append([
            inv.get("id"),
            inv.get("invoice_number"),
            str(inv.get("date") or ""),
            inv.get("type"),
            inv.get("customer_name"),
            inv.get("currency"),
            inv.get("taxable_value"),
            inv.get("tax_amount"),
            inv.get("total"),
            str(inv.get("items") or "")
        ])
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
